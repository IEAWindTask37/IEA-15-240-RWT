import os
try:
    import ruamel.yaml as yaml
except:
    try:
        import ruamel_yaml as yaml
    except:
        raise ImportError('No YAML package found')
import pandas as pd
import csv
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import matplotlib.cm as cm
from matplotlib.patches import Rectangle
from scipy.interpolate import interp1d

def find_nearest(array, value):
    return (np.abs(array - value)).argmin() 

def beamdyn_load(fname):
    f = open(fname, 'r')
    while True:
        line = f.readline()
        if line.find('station_total') >= 0:
            npts = int(line.split()[0])
        elif line.find('DISTRIBUTED PROPERTIES') >= 0:
            break

    r = np.zeros(npts)
    M = np.zeros( (6,6,npts) )
    K = np.zeros( (6,6,npts) )

    def read6x6(f):
        A = np.zeros((6,6))
        for k in range(6):
            line = f.readline()
            try:
                A[k,:] = np.fromstring(line, count=6, sep=' ')
            except:
                print(k,line)
                breakpoint()
        return A
        
    for ir in range(npts):
        r[ir] = float( f.readline().split()[0] )
        K[:,:,ir] = read6x6(f)
        f.readline()
        M[:,:,ir] = read6x6(f)
        f.readline()
        
    f.close()

    return M, K, r
        
def vabs_load(fname):
    with open(fname, 'r') as f:
        Alist = list( csv.reader(f) )

    A  = np.zeros( (6,6,1) )
    Ai = np.zeros( (6,6) )
    r  = np.array( [] )
    k  = 0
    for line in Alist:
        if len(line) == 0 or line[0].strip()=='': continue

        elif line[0].find('section') >= 0:
            if k > 0:
                # Finish last section
                A  = np.concatenate((A, Ai[:,:,np.newaxis]), axis=2)

            # Starting new section
            r  = np.r_[r, float(line[1])]
            Ai = np.zeros( (6,6) )
            k  = 0

        else:
            Ai[k,:] = np.array(line, dtype=np.float)
            k += 1
            
    return A, r


class RWT_Tabular(object):
    def __init__(self, finput, towDF=None, rotDF=None, layerDF=None, nacDF=None, overview=None):
        
        # Read ontology file into dictionary-like data structure
        f = open(finput, 'r')
        self.yaml = yaml.safe_load( f )
        f.close()
        
        # Store output file name
        folder_output = os.getcwd() + os.sep + 'outputs'
        _, fname      = os.path.split(finput)
        froot, _      = os.path.splitext( fname )
        self.fout     = folder_output + os.sep + froot + '_tabular.xlsx'

        # If provided, store blade, tower, and rotor data
        self.towDF   = towDF
        self.rotDF   = rotDF
        self.layerDF = layerDF
        self.nacDF   = nacDF
        self.overview = overview

        # Initialize workbook object
        self.wb   = Workbook()

        # Keep track of which airfoils to print out
        self.airfoil_list = []
        self.airfoil_span = []

        
    def write_all(self):
        self.write_overview()
        self.write_blade_outer()
        self.write_airfoils()
        self.write_blade_inner()
        self.write_blade_struct()
        self.write_tower_monopile()
        self.write_materials()
        self.write_rotor_performance()
        self.write_nacelle()
        self.cleanup()

    def write_overview(self):
        # Sheet name
        ws = self.wb.create_sheet(title = 'Overview')
        ws['A1'] = 'Parameter'
        ws['B1'] = 'IEA 15 MW Reference Wind Turbine'

        irow = 2
        for k,v in self.overview.items():
            myv = float(v) if type(v)==type(np.array([])) else v
            ws.cell(row=irow, column=1, value=k)
            ws.cell(row=irow, column=2, value=myv)
            irow += 1
        
        # Header row style formatting
        for cell in ws["1:1"]:
            cell.style = 'Headline 2'

        
    def write_airfoils(self):
        # Airfoil tables with plotting
        figsize = (6.5, 4)
        figCLA  = plt.figure(figsize=figsize)
        figLDA  = plt.figure(figsize=figsize)
        figCLCD = plt.figure(figsize=figsize)
        figAIR  = plt.figure(figsize=(8,4))
        axCLA   = figCLA.add_subplot(111)
        axLDA   = figLDA.add_subplot(111)
        axCLCD  = figCLCD.add_subplot(111)
        axAIR   = figAIR.add_subplot(111)
        labels0 = []
        labels  = []
        
        # Loop over airfoils, one tab per airfoil, append to plots
        for iaf in range(len(self.yaml['airfoils'])):
            if not self.yaml['airfoils'][iaf]['name'] in self.airfoil_list: continue
            
            # Sheet name will be airfoil name
            ws = self.wb.create_sheet(title = 'Airfoil '+self.yaml['airfoils'][iaf]['name'])

            # Write name and aerodynamic center
            ws['A1'] = 'name'
            ws['B1'] = self.yaml['airfoils'][iaf]['name']

            ws['A2'] = 'relative_thickness'
            ws['B2'] = self.yaml['airfoils'][iaf]['relative_thickness']

            ws['A3'] = 'aerodynamic_center'
            ws['B3'] = self.yaml['airfoils'][iaf]['aerodynamic_center']

            # Write airfoil xy-coordinates
            ws['A5'] = 'x'
            ws['B5'] = 'y'
            nxy      = len(self.yaml['airfoils'][iaf]['coordinates']['x'])
            for k in range(nxy):
                ws.cell(row=k+6, column=1, value=self.yaml['airfoils'][iaf]['coordinates']['x'][k])
                ws.cell(row=k+6, column=2, value=self.yaml['airfoils'][iaf]['coordinates']['y'][k])

            if not self.yaml['airfoils'][iaf]['name'].lower() == 'circular':
                axAIR.plot(self.yaml['airfoils'][iaf]['coordinates']['x'],
                           self.yaml['airfoils'][iaf]['coordinates']['y'])
                labels0.append(self.yaml['airfoils'][iaf]['name'])

            # Write airfoil polars
            row_start = 6+nxy+2
            for ipol in range(len(self.yaml['airfoils'][iaf]['polars'])):
                # Reynolds number label
                ws.cell(row=row_start, column=1, value='Reynolds')
                ws.cell(row=row_start, column=2, value=self.yaml['airfoils'][iaf]['polars'][ipol]['re'])

                # Create data array in pandas
                polmat = np.c_[self.yaml['airfoils'][iaf]['polars'][ipol]['c_l']['grid'],
                               self.yaml['airfoils'][iaf]['polars'][ipol]['c_l']['values'],
                               self.yaml['airfoils'][iaf]['polars'][ipol]['c_d']['values'],
                               self.yaml['airfoils'][iaf]['polars'][ipol]['c_m']['values']]
                polDF = pd.DataFrame(polmat, columns=['alpha [rad]',
                                                      'c_l',
                                                      'c_d',
                                                      'c_m'])

                # Append to worksheet
                for r in dataframe_to_rows(polDF, index=False, header=True):
                    ws.append(r)

                # Header row style formatting
                for cell in ws[str(row_start)+':'+str(row_start)]:
                    cell.style = 'Headline 1'
                for cell in ws[str(row_start+1)+':'+str(row_start+1)]:
                    cell.style = 'Headline 2'
                    
                # Prep for next polar
                row_start += polmat.shape[0]+1

                # Make plots with this airfoil
                if self.yaml['airfoils'][iaf]['name'] == 'circular': continue
                ReStr = np.format_float_scientific(self.yaml['airfoils'][iaf]['polars'][ipol]['re'], exp_digits=1, trim='-', sign=False)
                labels.append(self.yaml['airfoils'][iaf]['name']+' Re='+ReStr.replace('+',''))
                alpha = np.rad2deg(polDF['alpha [rad]'])
                ind = np.where(np.abs(alpha) <= 20.0)[0]
                axCLA.plot(alpha[ind], polDF['c_l'][ind], linewidth=1)
                axLDA.plot(alpha[ind], polDF['c_l'][ind]/polDF['c_d'][ind], linewidth=1)
                axCLCD.plot(polDF['c_d'][ind], polDF['c_l'][ind], linewidth=1)

            # Header row style formatting
            for cell in ws["1:1"]:
                cell.style = 'Headline 2'
            for cell in ws["5:5"]:
                cell.style = 'Headline 2'
        

        # Polish off plots
        axCLA.set_xlabel('Angle of Attack [deg]', fontsize=14, fontweight='bold')
        axCLA.set_ylabel('Lift coefficient, $c_l$', fontsize=14, fontweight='bold')
        axLDA.set_xlabel('Angle of Attack [deg]', fontsize=14, fontweight='bold')
        axLDA.set_ylabel('Lift over drag, $c_l/c_d$', fontsize=14, fontweight='bold')
        axCLCD.set_xlabel('Drag coefficient, $c_d$', fontsize=14, fontweight='bold')
        axCLCD.set_ylabel('Lift coefficient, $c_l$', fontsize=14, fontweight='bold')
        for ax in [axCLA, axLDA, axCLCD, axAIR]:
            plt.sca(ax)
            plt.xticks(fontsize=12)
            plt.yticks(fontsize=12)
            ax.grid(color=[0.8,0.8,0.8], linestyle='--')
            ax.legend(labels, loc='upper center', bbox_to_anchor=(1.25, 0.9), shadow=True, ncol=1)
        axAIR.legend(labels0, loc='lower left', bbox_to_anchor=(-0.1, -0.3), shadow=True, ncol=4)
        axAIR.axis('equal')
        figCLA.subplots_adjust(bottom = 0.15, left = 0.15)
        figLDA.subplots_adjust(bottom = 0.15, left = 0.15)
        figCLCD.subplots_adjust(bottom = 0.15, left = 0.15)
        figAIR.subplots_adjust(bottom = 0.15, left = 0.15)
        figCLA.savefig('outputs' + os.sep + 'airfoil_data-cl_alpha.pdf', pad_inches=0.1, bbox_inches='tight')
        figLDA.savefig('outputs' + os.sep + 'airfoil_data-clcd_alpha.pdf', pad_inches=0.1, bbox_inches='tight')
        figCLCD.savefig('outputs' + os.sep + 'airfoil_data-cl_cd.pdf', pad_inches=0.1, bbox_inches='tight')
        figAIR.savefig('outputs' + os.sep + 'airfoil_family.pdf', pad_inches=0.1, bbox_inches='tight')
        plt.close()
        
    def write_blade_outer(self):
        # Sheet name
        ws = self.wb.create_sheet(title = 'Blade Geometry')

        # Write airfoil xy-coordinates
        ws['A1'] = 'Spanwise position [r/R]'
        ws['B1'] = 'Airfoil name'
        npts = len(self.yaml['components']['blade']['outer_shape_bem']['airfoil_position']['grid'])
        for k in range(npts):
            ws.cell(row=k+2, column=1, value=self.yaml['components']['blade']['outer_shape_bem']['airfoil_position']['grid'][k])
            ws.cell(row=k+2, column=2, value=self.yaml['components']['blade']['outer_shape_bem']['airfoil_position']['labels'][k])
            self.airfoil_list.append( self.yaml['components']['blade']['outer_shape_bem']['airfoil_position']['labels'][k] )
            self.airfoil_span.append( self.yaml['components']['blade']['outer_shape_bem']['airfoil_position']['grid'][k] )

        ws.cell(row=npts+3, column=1, value='Profile')

        # Interpolation function for arbitraty grids
        mygrid = self.yaml['components']['blade']['outer_shape_bem']['chord']['grid']
        def myinterp(xgrid, val):
            try:
                y = interp1d(xgrid, val, kind='cubic', bounds_error=False, fill_value=0.0, assume_sorted=True).__call__(mygrid)
            except:
                y = interp1d(xgrid, val, kind='linear', bounds_error=False, fill_value=0.0, assume_sorted=True).__call__(mygrid)
            return y
        
        # Create blade geometry array in pandas
        geommat = np.c_[self.yaml['components']['blade']['outer_shape_bem']['chord']['grid'],
                        self.yaml['components']['blade']['outer_shape_bem']['chord']['values'],
                        myinterp(self.yaml['components']['blade']['outer_shape_bem']['twist']['grid'],
                                 self.yaml['components']['blade']['outer_shape_bem']['twist']['values']),
                        myinterp(self.yaml['components']['blade']['outer_shape_bem']['pitch_axis']['grid'],
                                 self.yaml['components']['blade']['outer_shape_bem']['pitch_axis']['values']),
                        myinterp(self.yaml['components']['blade']['outer_shape_bem']['reference_axis']['z']['grid'],
                                 self.yaml['components']['blade']['outer_shape_bem']['reference_axis']['z']['values']),
                        myinterp(self.yaml['components']['blade']['outer_shape_bem']['reference_axis']['x']['grid'],
                                 self.yaml['components']['blade']['outer_shape_bem']['reference_axis']['x']['values']),
                        myinterp(self.yaml['components']['blade']['outer_shape_bem']['reference_axis']['y']['grid'],
                                 self.yaml['components']['blade']['outer_shape_bem']['reference_axis']['y']['values'])  ]
        geomDF = pd.DataFrame(geommat, columns=['Spanwise position [r/R]',
                                              'Chord [m]',
                                              'Twist [rad]',
                                              'Pitch axis [x/c]',
                                              'Span [m]',
                                              'Prebend [m]',
                                              'Sweep [m]'])
        
        # Set convention that positive prebend means upward tilt to give tower strike margin
        geomDF['Prebend [m]'] *= -1.0
        
        # Append to worksheet
        for r in dataframe_to_rows(geomDF, index=False, header=True):
            ws.append(r)

        # Header row style formatting
        for cell in ws["1:1"]:
            cell.style = 'Headline 2'
        for cell in ws[str(npts+3)+':'+str(npts+3)]:
            cell.style = 'Headline 1'
        for cell in ws[str(npts+4)+':'+str(npts+4)]:
            cell.style = 'Headline 2'

        
    def write_blade_inner(self):
        # Set number of entries
        naf  = len(self.airfoil_span)
        nweb = len(self.yaml['components']['blade']['internal_structure_2d_fem']['webs'])
        nlay = len(self.yaml['components']['blade']['internal_structure_2d_fem']['layers'])

        # Pre-loop to set grid, names, materials, etc
        mygrid  = np.array(self.airfoil_span)
        matlist = []
        weblist = []
        for k in range(nweb):
            for ikey in self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k].keys():
                if ikey in ['name']:
                    weblist.append( self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k][ikey] )
                elif ikey == 'material':
                    matlist.append( self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k][ikey] )
                else:
                    mygrid = np.r_[mygrid, self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k][ikey]['grid']]

        for k in range(nlay):
            for ikey in self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k].keys():
                if ikey in ['name', 'midpoint_nd_arc','start_nd_arc','end_nd_arc','side','web']:
                    continue
                elif ikey == 'material':
                    matlist.append( self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k][ikey] )
                else:
                    mygrid = np.r_[mygrid, self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k][ikey]['grid']]
        mygrid  = np.unique( mygrid )

        # Get unique list of materials and associate with unique color
        matlist = sorted(list(set(matlist)))
        colors = cm.Dark2( np.linspace(0, 1, len(matlist)) ).tolist()
        matcolor = dict(zip(matlist, colors))
        matcolor['CarbonUD'] = [0.0, 0.0, 0.0, 1.0]
        leglist = [Rectangle((0, 0), 1, 1, color=matcolor[m]) for m in matlist]
        
        # Interpolation function for arbitraty layers and webs
        def myinterp(xgrid, val):
            try:
                y = interp1d(xgrid, val, kind='cubic', bounds_error=False, fill_value=0.0, assume_sorted=True).__call__(mygrid)
            except:
                y = interp1d(xgrid, val, kind='linear', bounds_error=False, fill_value=0.0, assume_sorted=True).__call__(mygrid)
                #y = np.interp(mygrid, xgrid, val, left=0.0, right=0.0)
            return y
        
        # Loop over layers and store layup data in master tables for plotting and writing to Excel table
        afstack  = [[] for m in range(naf)]
        webstack = [[] for m in range(naf)]
        for k in range(nlay):
            lname = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['name']
            lmat  = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['material']

            iDF = self.layerDF[k]
            ithick = myinterp(iDF['Span'], iDF['Thickness [m]'])
            idir   = myinterp(iDF['Span'], iDF['Fiber angle [deg]'])

            if lname.lower().find('web') >= 0:
                iweb   = weblist.index( self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['web'] )

                for iaf in range(naf):
                    ind = np.where(mygrid == self.airfoil_span[iaf])[0][0]
                    webstack[iaf].append([lname, lmat, iweb, 1e3*ithick[ind], idir[ind]] )
                    
            else:
                lay_beg = myinterp(iDF['Span'], iDF['Layer Start'])
                lay_end = myinterp(iDF['Span'], iDF['Layer End'])

                for iaf in range(naf):
                    ind = np.where(mygrid == self.airfoil_span[iaf])[0][0]
                    afstack[iaf].append( [lname, lmat, lay_beg[ind], lay_end[ind], 1e3*ithick[ind], idir[ind]] )

        # Area plot for web.  Could probably do a bar chart, but will do fill_between for consistency
        # Initialize the plotting
        x     = np.linspace(0.0, float(nweb+1), 1000)
        fig   = plt.figure(figsize=(8,4))
        ax    = fig.add_subplot(111)
        yBase = np.zeros( x.shape )
        for iaf in range(naf):
            nstack = len(webstack[iaf])

            for k in range(nstack):
                # Skip if no thickness
                if webstack[iaf][k][3] == 0.0: continue

                # Thickness adder for area plot
                ywhere = np.array( [False]*x.size )
                iweb   = webstack[iaf][k][2]
                ywhere[np.logical_and(x >= np.round(iweb*1.1+0.1*iaf+0.5,1), x < np.round(iweb*1.1+0.1*iaf+0.6,1))] = True
                yadd   = np.zeros( x.shape )
                yadd[ywhere] = webstack[iaf][k][3]
                yPlot  = yBase + yadd
                
                # Fill in the area and remember the color and label
                ax.fill_between(x, yBase, yPlot, where=ywhere, step='mid', fc=matcolor[ webstack[iaf][k][1] ], ec='face')

                # Increment the baseline for the next layer
                yBase = yPlot 

        # Clean-up the plotting
        vy = ax.get_ylim()
        ax.axis([0.4, 1.1*nweb+0.4, 0.0, vy[1]+20])
        leg = ax.legend(leglist, matlist, loc = 'upper left', ncol=3, bbox_to_anchor = (0.0, 1.0))
        #ax.set_xlabel(weblist[0]+'             '+weblist[1]+'             '+weblist[2], size=14, weight='bold')
        ax.set_xlabel(weblist[0]+'                                       '+weblist[1], size=14, weight='bold')
        ax.set_ylabel('Thickness [mm]', size=14, weight='bold')
        vy = ax.get_ylim()
        xtick = 0.5 + 0.1*np.arange(len(self.airfoil_list)) + 0.05
        labs  = [self.airfoil_list[m]+' '+str(int(np.round(1e2*self.airfoil_span[m])))+'%' for m in range(naf)]
        #ax.set_xticks( np.r_[xtick, xtick+1, xtick+2] )
        ax.set_xticks( np.r_[xtick, xtick+1.1] )
        #ax.set_xticklabels( labs+labs+labs, rotation='vertical' )
        ax.set_xticklabels( labs+labs, rotation='vertical' )
        fig.subplots_adjust(bottom = 0.15, left = 0.15)
        fig.savefig('outputs' + os.sep + 'web_layup.pdf', pad_inches=0.1, bbox_inches='tight')
                
            
        # Area plots for airfoil skin
        x = np.linspace(0.0, 1.0, 1000)
        for iaf in range(naf):
            # Initialize the plotting
            fig    = plt.figure(figsize=(12,4))
            ax     = fig.add_subplot(111)
            yBase  = np.zeros( x.shape )
            nstack = len(afstack[iaf])

            for k in range(nstack):
                # Skip if no thickness
                if afstack[iaf][k][4] == 0.0: continue
                
                # S-coord region for this layer
                ibeg = find_nearest(x, afstack[iaf][k][2])
                iend = find_nearest(x, afstack[iaf][k][3])

                # Thickness adder for area plot
                ywhere = np.array( [False]*x.size )
                if ibeg < iend:
                    ywhere[ibeg:iend] = True
                elif ibeg > iend:
                    ywhere[ibeg:] = True
                    ywhere[:iend] = True
                else:
                    ywhere[ibeg] = True
                    
                yadd  = np.zeros( x.shape )
                yadd[ywhere] = afstack[iaf][k][4]
                yPlot = yBase + yadd

                # Fill in the area and remember the color and label
                ax.fill_between(x, yBase, yPlot, where=ywhere, step='mid', fc=matcolor[ afstack[iaf][k][1] ], ec='face')

                # Increment the baseline for the next layer
                yBase = yPlot 
                
            # Clean-up the plotting
            vy = ax.get_ylim()
            ax.axis([0.0, 1.0, 0.0, vy[1]+30])
            leg = ax.legend(leglist, matlist, loc = 'upper left', ncol=3, bbox_to_anchor = (0.0, 1.0))
            ax.set_xlabel('Airfoil s-coordinate [-]', size=14, weight='bold')
            ax.set_ylabel('Thickness [mm]', size=14, weight='bold')
            vy = ax.get_ylim()
            ax.text(0.6, 0.95*np.diff(vy), self.airfoil_list[iaf]+' airfoil, '+str(int(np.round(1e2*self.airfoil_span[iaf])))+'% span',
                    size=12, weight='bold')
            ax.set_xticks([0.0, 0.25, 0.5, 0.75, 1.0])
            ax.set_xticklabels(['TE','Suction Side', 'LE','Pressure Side','TE'])
            fig.subplots_adjust(bottom = 0.15, left = 0.15)
            fig.savefig('outputs' + os.sep + 'layers_'+self.airfoil_list[iaf]+'_'+str(int(np.round(1e2*self.airfoil_span[iaf])))+'.pdf',
                        pad_inches=0.1, bbox_inches='tight')
            plt.close()
        
        # Grab shear web data for excel sheet
        web3_grid, web3_ss, web3_ps = None, None, None
        for k in range(nweb):
            if self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k]['name'] == 'web0':
                web1_grid = self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k]['start_nd_arc']['grid']
                web1_ss   = self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k]['start_nd_arc']['values']
                web1_ps   = self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k]['end_nd_arc']['values']
            elif self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k]['name'] == 'web1':
                web2_grid = self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k]['start_nd_arc']['grid']
                web2_ss   = self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k]['start_nd_arc']['values']
                web2_ps   = self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k]['end_nd_arc']['values']
            elif self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k]['name'] == 'web2':
                web3_grid = self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k]['start_nd_arc']['grid']
                web3_ss   = self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k]['start_nd_arc']['values']
                web3_ps   = self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k]['end_nd_arc']['values']
            else:
                print('Unknown web, ',self.yaml['components']['blade']['internal_structure_2d_fem']['webs'][k]['name'])

        # Grab spar cap and LE/TE reinforcement data for excel sheet
        for k in range(nlay):
            if self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['name'].lower() == 'spar_cap_ss':
                sparcap_ss_grid = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['start_nd_arc']['grid']
                sparcap_ss_th   = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['thickness']['values']
                sparcap_ss_wgrid = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['width']['grid']
                sparcap_ss_wid  = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['width']['values']
                sparcap_ss_beg  = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['start_nd_arc']['values']
                sparcap_ss_end  = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['end_nd_arc']['values']
            elif self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['name'].lower() == 'spar_cap_ps':
                sparcap_ps_grid = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['start_nd_arc']['grid']
                sparcap_ps_th   = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['thickness']['values']
                sparcap_ps_wgrid = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['width']['grid']
                sparcap_ps_wid  = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['width']['values']
                sparcap_ps_beg  = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['start_nd_arc']['values']
                sparcap_ps_end  = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['end_nd_arc']['values']
            elif self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['name'] == 'LE_reinforcement':
                reinf_le_thgrid = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['thickness']['grid']
                reinf_le_th     = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['thickness']['values']
                reinf_le_wgrid  = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['width']['grid']
                reinf_le_wid    = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['width']['values']
            elif self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['name'].lower() == 'te_reinforcement_ss':
                reinf_te_ss_thgrid = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['thickness']['grid']
                reinf_te_ss_th     = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['thickness']['values']
                reinf_te_ss_wgrid  = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['width']['grid']
                reinf_te_ss_wid    = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['width']['values']
            elif self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['name'].lower() == 'te_reinforcement_ps':
                reinf_te_ps_thgrid = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['thickness']['grid']
                reinf_te_ps_th     = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['thickness']['values']
                reinf_te_ps_wgrid  = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['width']['grid']
                reinf_te_ps_wid    = self.yaml['components']['blade']['internal_structure_2d_fem']['layers'][k]['width']['values']
            else:
                continue

        # Put structural overview in a DF
        crossDF = pd.DataFrame()
        crossDF['Span position'] = mygrid
        crossDF['Shear web-1 SS s-coord'] = myinterp(web1_grid, web1_ss)
        crossDF['Shear web-1 PS s-coord'] = myinterp(web1_grid, web1_ps)
        crossDF['Shear web-2 SS s-coord'] = myinterp(web2_grid, web2_ss)
        crossDF['Shear web-2 PS s-coord'] = myinterp(web2_grid, web2_ps)
        if not web3_grid is None:
            crossDF['Shear web-3 SS s-coord'] = myinterp(web3_grid, web3_ss)
            crossDF['Shear web-3 PS s-coord'] = myinterp(web3_grid, web3_ps)
        crossDF['Spar cap SS begin s-coord'] = myinterp(sparcap_ss_grid, sparcap_ss_beg)
        crossDF['Spar cap SS width [m]']     = myinterp(sparcap_ss_wgrid, sparcap_ss_wid)
        crossDF['Spar cap SS thick [m]']     = myinterp(sparcap_ss_grid, sparcap_ss_th )
        crossDF['Spar cap PS begin s-coord'] = myinterp(sparcap_ps_grid, sparcap_ps_beg)
        crossDF['Spar cap PS width [m]']     = myinterp(sparcap_ps_wgrid, sparcap_ps_wid)
        crossDF['Spar cap PS thick [m]']     = myinterp(sparcap_ps_grid, sparcap_ps_th )
        crossDF['LE reinf width [m]']        = myinterp(reinf_le_wgrid, reinf_le_wid)
        crossDF['LE reinf thick [m]']        = myinterp(reinf_le_thgrid, reinf_le_th )
        crossDF['TE reinf SS width [m]']     = myinterp(reinf_te_ss_wgrid, reinf_te_ss_wid)
        crossDF['TE reinf SS thick [m]']     = myinterp(reinf_te_ss_thgrid, reinf_te_ss_th )
        crossDF['TE reinf PS width [m]']     = myinterp(reinf_te_ps_wgrid, reinf_te_ps_wid)
        crossDF['TE reinf PS thick [m]']     = myinterp(reinf_te_ps_thgrid, reinf_te_ps_th )

        # Cross section over-view sheet
        ws = self.wb.create_sheet(title = 'Blade Support Structure')
        for r in dataframe_to_rows(crossDF, index=False, header=True):
            ws.append(r)
        for cell in ws["1:1"]:
            cell.style = 'Headline 2'

        # Lay-up sheet for airfoil skin 
        ws = self.wb.create_sheet(title = 'Blade Shell Layup')
        irow = 1
        for iaf in range(naf):
            icell = ws.cell(row=irow, column=1, value='Airfoil')
            icell.style = 'Headline 1'
            icell = ws.cell(row=irow+1, column=1, value='Pct Span')
            icell.style = 'Headline 1'
            icell = ws.cell(row=irow, column=2, value=self.airfoil_list[iaf])
            icell = ws.cell(row=irow+1, column=2, value=str(1e2*self.airfoil_span[iaf]))
            icell.style = 'Percent'

            myDF = pd.DataFrame( data=afstack[iaf], columns=['Layer Name',
                                                             'Material',
                                                             'S-coord start',
                                                             'S-coord end',
                                                             'Thickness [mm]',
                                                             'Fiber direction [?]'])
            nstack = len(afstack[iaf])
            for r in dataframe_to_rows(myDF, index=False, header=True):
                ws.append(r)
            for cell in ws[str(irow+2)+':'+str(irow+2)]:
                cell.style = 'Headline 2'
            irow += nstack+5
                         
        # Lay-up sheet for shear webs
        ws = self.wb.create_sheet(title = 'Shear Web Layup')
        irow = 1
        for iaf in range(naf):
            icell = ws.cell(row=irow, column=1, value='Airfoil')
            icell.style = 'Headline 1'
            icell = ws.cell(row=irow+1, column=1, value='Pct Span')
            icell.style = 'Headline 1'
            icell = ws.cell(row=irow, column=2, value=self.airfoil_list[iaf])
            icell = ws.cell(row=irow+1, column=2, value=str(1e2*self.airfoil_span[iaf]))
            icell.style = 'Percent'

            myDF = pd.DataFrame( data=webstack[iaf], columns=['Layer Name',
                                                             'Material',
                                                             'iweb',
                                                             'Thickness [mm]',
                                                             'Fiber direction [?]'])
            myDF.sort_values('iweb', inplace=True)
            myDF['Shear web'] = [weblist[m] for m in myDF['iweb']]
            myDF = myDF[['Shear web','Layer Name','Material','Thickness [mm]','Fiber direction [?]']]
            myDF = myDF.loc[myDF['Thickness [mm]'] > 0.0].copy()
            
            nstack = len(myDF.index)
            for r in dataframe_to_rows(myDF, index=False, header=True):
                ws.append(r)
            for cell in ws[str(irow+2)+':'+str(irow+2)]:
                cell.style = 'Headline 2'
            irow += nstack+5
                         
            
    def write_blade_struct(self):
        
        # Load in VABS data
        #froot  = os.path.join('..','OpenFAST','IEA-15-240-RWT','VABS','IEA-15-240-RWT_vabs_beam_properties_')
        #fnames = ['mass_matrices.csv','stiff_matrices.csv','general.csv']
        #M, Mr  = vabs_load(froot+fnames[0])
        #K, Kr  = vabs_load(froot+fnames[1])
        #sumDF = pd.read_csv(froot+fnames[2], header=1, index_col=0)

        fbeamdyn = os.path.join('..','OpenFAST','IEA-15-240-RWT','IEA-15-240-RWT_BeamDyn_blade.dat')
        M, K, r = beamdyn_load( fbeamdyn )
        
        mydata = np.c_[M[0,0,:], M[0,1,:], M[0,2,:], M[0,3,:], M[0,4,:], M[0,5,:], 
                       M[1,1,:], M[1,2,:], M[1,3,:], M[1,4,:], M[1,5,:], 
                       M[2,2,:], M[2,3,:], M[2,4,:], M[2,5,:], 
                       M[3,3,:], M[3,4,:], M[3,5,:], 
                       M[4,4,:], M[4,5,:], 
                       M[5,5,:], 
                       K[0,0,:], K[0,1,:], K[0,2,:], K[0,3,:], K[0,4,:], K[0,5,:], 
                       K[1,1,:], K[1,2,:], K[1,3,:], K[1,4,:], K[1,5,:], 
                       K[2,2,:], K[2,3,:], K[2,4,:], K[2,5,:], 
                       K[3,3,:], K[3,4,:], K[3,5,:], 
                       K[4,4,:], K[4,5,:], 
                       K[5,5,:]]

        labels = ['M_11','M_12','M_13','M_14','M_15','M_16',
                  'M_22','M_23','M_24','M_25','M_26',
                  'M_33','M_34','M_35','M_36',
                  'M_44','M_45','M_46',
                  'M_55','M_56',
                  'M_66',
                  'K_11','K_12','K_13','K_14','K_15','K_16',
                  'K_22','K_23','K_24','K_25','K_26',
                  'K_33','K_34','K_35','K_36',
                  'K_44','K_45','K_46',
                  'K_55','K_56',
                  'K_66']

        #bladeStructDF = pd.DataFrame(data=mydata, columns=labels, index=Mr)
        bladeStructDF = pd.DataFrame(data=mydata, columns=labels, index=r)
        
        # Write to sheet
        ws = self.wb.create_sheet(title = 'Blade Structural Properties')
        ws['A1'] = 'BeamDyn Coordinate System (see https://wind.nrel.gov/nwtc/docs/BeamDyn_Manual.pdf)'
        #for r in dataframe_to_rows(sumDF, index=True, header=True):
        #    ws.append(r)
        #ws['A16'] = ''
        for r in dataframe_to_rows(bladeStructDF, index=True, header=True):
            ws.append(r)
        
        # Header row style formatting
        for cell in ws["2:2"]:
            cell.style = 'Headline 2'
        for cell in ws["17:17"]:
            cell.style = 'Headline 2'

        
        '''
        # Make plot
        blade_x  = np.array(self.yaml['components']['blade']['outer_shape_bem']['chord']['grid'])
        def myinterp(xgrid, val):
            return interp1d(xgrid, val, kind='cubic', bounds_error=False, fill_value=0.0, assume_sorted=True).__call__(blade_x)
        blade_le = (np.array(self.yaml['components']['blade']['outer_shape_bem']['chord']['values']) *
                    myinterp(self.yaml['components']['blade']['outer_shape_bem']['pitch_axis']['grid'],
                             self.yaml['components']['blade']['outer_shape_bem']['pitch_axis']['values']) )
        blade_te = (np.array(self.yaml['components']['blade']['outer_shape_bem']['chord']['values']) *
                    (1. - myinterp(self.yaml['components']['blade']['outer_shape_bem']['pitch_axis']['grid'],
                                   self.yaml['components']['blade']['outer_shape_bem']['pitch_axis']['values']) ) )
        xx      = sumDF.index 
        y_mass  = sumDF['Mass center (chordwise), m']
        y_neut  = sumDF['Neutral axes (chordwise), m']
        y_geo   = sumDF['Geometric center (chordwise), m']
        y_shear = sumDF['Shear center (chordwise), m']
        fig = plt.figure(figsize=(8,4))
        ax  = fig.add_subplot(111)
        ax.plot(blade_x, np.zeros_like(blade_x), 'k:')
        ax.plot(np.array(xx), np.array(y_mass), linewidth=2)
        ax.plot(np.array(xx), np.array(y_neut), linewidth=2)
        ax.plot(np.array(xx), np.array(y_geo), linewidth=2)
        ax.plot(np.array(xx), np.array(y_shear), linewidth=2)
        ax.plot(np.array(blade_x), np.array(blade_le), 'k', np.array(blade_x), np.array(-blade_te), 'k', linewidth=2.5)
        ax.legend(['Pitch axis','Mass center','Neutral center','Geometric center','Shear center'])
        ax.set_xlabel('Blade span r/R', size=14, weight='bold')
        ax.set_ylabel('Chordwise [m]', size=14, weight='bold')
        fig.savefig('outputs' + os.sep + 'planform.pdf', pad_inches=0.1, bbox_inches='tight')
        '''

        
    def write_tower_monopile(self):
        if not self.towDF is None:
            ws = self.wb.create_sheet(title = 'Tower Properties')
            for r in dataframe_to_rows(self.towDF, index=False, header=True):
                ws.append(r)
        
            # Header row style formatting
            for cell in ws["1:1"]:
                cell.style = 'Headline 2'
            
    
    def write_materials(self):
        # Sheet name
        ws = self.wb.create_sheet(title = 'Material Properties')

        # Initialize containers
        nmat = len(self.yaml['materials'])
        matname = [''] * nmat
        E       = np.zeros( (nmat,3) )
        G       = np.zeros( (nmat,3) )
        nu      = np.zeros( (nmat,3) )
        rho     = np.zeros( (nmat,)  )
        Xt      = np.zeros( (nmat,3) )
        Xc      = np.zeros( (nmat,3) )
        fvf     = np.zeros( (nmat,)  )
        desc    = [''] * nmat
        srcs    = [''] * nmat

        # Loop over all materials and append
        for k in range(len(self.yaml['materials'])):
            matname[k] = self.yaml['materials'][k]['name']
            rho[k]     = self.yaml['materials'][k]['rho']

            if 'fvf' in self.yaml['materials'][k].keys():
                fvf[k] = self.yaml['materials'][k]['fvf']

            if 'source' in self.yaml['materials'][k].keys():
                srcs[k] = self.yaml['materials'][k]['source']

            if 'description' in self.yaml['materials'][k].keys():
                desc[k] = self.yaml['materials'][k]['description']

            if 'E' in self.yaml['materials'][k].keys():
                temp = self.yaml['materials'][k]['E']
                if isinstance(temp, list) and len(temp) == 3:
                    E[k,:] = temp
                else:
                    E[k,:] = temp*np.ones(3)

            if 'G' in self.yaml['materials'][k].keys():
                temp = self.yaml['materials'][k]['G'] 
                if isinstance(temp, list) and len(temp) == 3:
                    G[k,:] = temp
                else:
                    G[k,:] = temp*np.ones(3)

            if 'nu' in self.yaml['materials'][k].keys():
                temp = self.yaml['materials'][k]['nu']
                if isinstance(temp, list) and len(temp) == 3:
                    nu[k,:] = temp
                else:
                    nu[k,:] = temp*np.ones(3)

            if 'Xt' in self.yaml['materials'][k].keys():
                temp = self.yaml['materials'][k]['Xt'] 
                if isinstance(temp, list) and len(temp) == 3:
                    Xt[k,:] = temp
                else:
                    Xt[k,:] = temp*np.ones(3)

            if 'Xc' in self.yaml['materials'][k].keys():
                temp = self.yaml['materials'][k]['Xc']
                if isinstance(temp, list) and len(temp) == 3:
                    Xc[k,:] = temp
                else:
                    Xc[k,:] = temp*np.ones(3)

        # Fill in void values
        fvf[fvf==0.0] = np.nan
        
        # Build up matrix
        matDF = pd.DataFrame()
        matDF['Material name'] = matname
        matDF['Density [kg/m^3]'] = rho
        matDF['Fiber vol fraction'] = fvf
        matDF['Young Modulus E_1 [MPa]'] = 1e-6*E[:,0]
        matDF['Young Modulus E_2 [MPa]'] = 1e-6*E[:,1]
        matDF['Young Modulus E_3 [MPa]'] = 1e-6*E[:,2]
        matDF['Shear Modulus G_1 [MPa]'] = 1e-6*G[:,0]
        matDF['Shear Modulus G_2 [MPa]'] = 1e-6*G[:,1]
        matDF['Shear Modulus G_3 [MPa]'] = 1e-6*G[:,2]
        matDF['Poisson ratio nu_1'] = nu[:,0]
        matDF['Poisson ratio nu_2'] = nu[:,1]
        matDF['Poisson ratio nu_3'] = nu[:,2]
        matDF['Tensile failure Xt_1 [MPa]'] = 1e-6*Xt[:,0]
        matDF['Tensile failure Xt_2 [MPa]'] = 1e-6*Xt[:,1]
        matDF['Tensile failure Xt_3 [MPa]'] = 1e-6*Xt[:,2]
        matDF['Compressive failure Xc_1 [MPa]'] = 1e-6*Xc[:,0]
        matDF['Compressive failure Xc_2 [MPa]'] = 1e-6*Xc[:,1]
        matDF['Compressive failure Xc_3 [MPa]'] = 1e-6*Xc[:,2]
        matDF['Description'] = desc
        matDF['Reference']   = srcs

        # Write it out
        for r in dataframe_to_rows(matDF, index=False, header=True):
            ws.append(r)
        
        # Header row style formatting
        for cell in ws["1:1"]:
            cell.style = 'Headline 2'
        
            
    
    def write_rotor_performance(self):
        ws = self.wb.create_sheet(title = 'Rotor Performance')

        # Use OpenFAST output if it is available
        foutput = '..'+os.sep+'OpenFAST'+os.sep+'outputs'+os.sep+'IEA-15-240-RWT_steady.yaml'
        if os.path.exists(foutput):
            f = open(foutput, 'r')
            fastout = yaml.safe_load( f )
            f.close()

            rotmat = np.c_[fastout['Wind1VelX']['mean'],
                           fastout['BldPitch1']['mean'],
                           1e-3*np.array(fastout['GenPwr']['mean']),
                           fastout['RotSpeed']['mean'],
                           120.0*np.array(fastout['RotSpeed']['mean'])*2*np.pi/60.0,
                           1e-3*np.array(fastout['RotThrust']['mean']),
                           1e-3*np.array(fastout['RotTorq']['mean']),
                           1e-3*np.array(fastout['GenTq']['mean']),
                           fastout['RtAeroCp']['mean'],
                           fastout['RtAeroCt']['mean']]
            cols = ['Wind [m/s]',
                    'Pitch [deg]',
                    'Power [MW]',
                    'Rotor Speed [rpm]',
                    'Tip Speed [m/s]',
                    'Rotor Thrust [MN]',
                    'Rotor Torque [MNm]',
                    'Generator Torque [MNm]',
                    'Rotor Cp [-]',
                    'Rotor Ct [-]']
            myDF = pd.DataFrame(rotmat, columns=cols)
            
        elif not self.rotDF is None:
            # Use WISDEM output
            myDF = self.rotDF

        else:
            return

        # Write to the file
        for r in dataframe_to_rows(myDF, index=False, header=True):
            ws.append(r)
        
        # Header row style formatting
        for cell in ws["1:1"]:
            cell.style = 'Headline 2'


    def write_nacelle(self):
        if not self.nacDF is None:
            ws = self.wb.create_sheet(title = 'Nacelle Mass Properties')
            for r in dataframe_to_rows(self.nacDF, index=True, header=True):
                ws.append(r)
        
            # Header row style formatting
            for cell in ws["1:1"]:
                cell.style = 'Headline 2'

        # Dump to latex file
        #with open('nac.tbl','w') as f:
        #    self.nacDF.to_latex(f, index=True)
        
        
    def cleanup(self):
        # Remove empty sheet
        self.wb.active = 0
        ws = self.wb.active
        self.wb.remove(ws)
        
        # Save workbook to file
        self.wb.save(self.fout)

        
if __name__ == '__main__':
    


    # Ontology file as input
    ontology_dir  = os.path.dirname( os.path.dirname( os.path.realpath(__file__)) ) + os.sep + 'WT_Ontology'
    fontology     = ontology_dir + os.sep + 'IEA-15-240-RWT_FineGrid.yaml'
    
    myobj = RWT_Tabular(fontology)
    myobj.write_all()
