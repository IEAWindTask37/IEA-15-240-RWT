from __future__ import print_function
import numpy as np
from scipy.interpolate import PchipInterpolator
import os, shutil, copy, time, sys
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from pprint import pprint
from openmdao.api import IndepVarComp, ExplicitComponent, Group, Problem, ScipyOptimizeDriver, SqliteRecorder, NonlinearRunOnce, DirectSolver, CaseReader
try:
    from openmdao.api import pyOptSparseDriver
except:
    pass
from wisdem.rotorse.rotor import RotorSE, Init_RotorSE_wRefBlade
from wisdem.rotorse.rotor_geometry_yaml import ReferenceBlade
from wisdem.commonse.turbine_constraints import TurbineConstraints
from wisdem.assemblies.fixed_bottom.monopile_assembly_turbine2 import MonopileTurbine, Init_MonopileTurbine
#from wisdem.commonse.mpi_tools import MPI

from wisdem.aeroelasticse.FAST_reader import InputReader_Common, InputReader_OpenFAST, InputReader_FAST7
#from wisdem.rotorse.rotor_visualization import plot_lofted



# Class to print outputs on screen
class Outputs_2_Screen(ExplicitComponent):
    def setup(self):
        
        # self.add_input('chord',                val=np.zeros(NPTS))
        # self.add_input('theta',                val=np.zeros(NPTS))
        self.add_input('bladeLength',          val=0.0, units = 'm')
        self.add_input('total_blade_cost',     val=0.0, units = 'USD')
        self.add_input('mass_one_blade',       val=0.0, units = 'kg')
        self.add_input('tower_mass',           val=0.0, units = 'kg')
        self.add_input('tower_cost',           val=0.0, units = 'USD')
        self.add_input('control_tsr',          val=0.0)
        self.add_input('AEP',                  val=0.0, units = 'GW * h')
        self.add_input('lcoe',                 val=0.0, units = 'USD/MW/h')
        self.add_input('rated_T',              val=0.0, units = 'MN')
        self.add_input('root_bending_moment',  val=0.0, units = 'MN * m')
        self.add_input('tip_deflection',       val=0.0, units = 'm')
        self.add_input('tip_deflection_ratio', val=0.0)

        
    def compute(self, inputs, outputs):
        print('########################################')
        print('Optimization variables')
        # print('Max chord:   {:8.3f} m'.format(max(inputs['chord'])))
        print('TSR:         {:8.3f} -'.format(inputs['control_tsr'][0]))
        print('')
        print('Constraints')
        print('Max TD:      {:8.3f} m'.format(inputs['tip_deflection'][0]))
        print('TD ratio:    {:8.10f} -'.format(inputs['tip_deflection_ratio'][0]))
        print('')
        print('Objectives')
        print('AEP:         {:8.10f} GWh'.format(inputs['AEP'][0]))
        print('Blade mass:  {:8.3f} kg'.format(inputs['mass_one_blade'][0]))
        print('Blade cost:  {:8.3f} $'.format(inputs['total_blade_cost'][0]))
        print('Tower mass:  {:8.3f} kg'.format(inputs['tower_mass'][0]))
        print('Tower cost:  {:8.3f} $'.format(inputs['tower_cost'][0]))
        print('LCoE:        {:8.3f} $/MWh'.format(inputs['lcoe'][0]))
        print('########################################')

class Convergence_Trends_Opt(ExplicitComponent):
    def initialize(self):
        
        self.options.declare('folder_output')
        self.options.declare('optimization_log')
        
    def compute(self, inputs, outputs):
        
        folder_output       = self.options['folder_output']
        optimization_log    = self.options['folder_output'] + self.options['optimization_log']

        if os.path.exists(optimization_log):
        
            cr = CaseReader(optimization_log)
            cases = cr.list_cases()
            rec_data = {}
            iterations = []
            for i, casei in enumerate(cases):
                iterations.append(i)
                it_data = cr.get_case(casei)
                
                # parameters = it_data.get_responses()
                for parameters in [it_data.get_responses(), it_data.get_design_vars()]:
                    for j, param in enumerate(parameters.keys()):
                        if i == 0:
                            rec_data[param] = []
                        rec_data[param].append(parameters[param])

            for param in rec_data.keys():
                fig, ax = plt.subplots(1,1,figsize=(5.3, 4))
                ax.plot(iterations, rec_data[param])
                ax.set(xlabel='Number of Iterations' , ylabel=param)
                fig_name = 'Convergence_trend_' + param + '.png'
                fig.savefig(folder_output + fig_name)
                plt.close(fig)

# Group to link the openmdao components
class Optimize_MonopileTurbine(Group):

    def initialize(self):
        self.options.declare('RefBlade')
        self.options.declare('folder_output' ,  default='')
        self.options.declare('FASTpref',        default={})
        self.options.declare('Nsection_Tow',    default = 6)
        self.options.declare('VerbosityCosts',  default = False)
        self.options.declare('user_update_routine',     default=None)
        
    def setup(self):
        RefBlade             = self.options['RefBlade']
        folder_output        = self.options['folder_output']
        Nsection_Tow         = self.options['Nsection_Tow']
        VerbosityCosts       = self.options['VerbosityCosts']
        user_update_routine  = self.options['user_update_routine']
        FASTpref             = self.options['FASTpref']
    
        self.add_subsystem('lb_wt', MonopileTurbine(RefBlade=blade, Nsection_Tow = Nsection_Tow, VerbosityCosts = VerbosityCosts, user_update_routine=user_update_routine, FASTpref=FASTpref), promotes=['*'])
        # Post-processing
        self.add_subsystem('outputs_2_screen',  Outputs_2_Screen(), promotes=['*'])
        self.add_subsystem('conv_plots',        Convergence_Trends_Opt(folder_output = folder_output, optimization_log = 'log_opt_' + RefBlade['config']['name']))
        

def set_web3_offset(blade):
    # User Routine to set the 3rd web offset to be at 90% chord for the NREL 15MW
    web_name     = 'third_web'
    web_position = 0.9
    # ------------------------

    web_idx  = [idx for idx, web in enumerate(blade['st']['webs']) if web['name']==web_name][0]

    r_in          = blade['ctrl_pts']['r_in']
    chord_in      = blade['ctrl_pts']['chord_in']
    p_le_spline   = PchipInterpolator(blade['pf']['s'], blade['pf']['p_le'])
    p_le_in       = p_le_spline(r_in)
    
    offset_in     = [chord_i*(web_position-p_le_i) for p_le_i, chord_i in zip(p_le_in, chord_in)]
    offset_spline = PchipInterpolator(r_in, offset_in)
    offset        = offset_spline(blade['pf']['s'])

    blade['st']['webs'][web_idx]['offset_x_pa']['grid']   = blade['pf']['s']
    blade['st']['webs'][web_idx]['offset_x_pa']['values'] = offset

    return blade

if __name__ == "__main__":
    optFlag       = False
    
    fname_schema  = 'IEAontology_schema.yaml'
    fname_input   = 'IEA-15-240-RWT.yaml'
    fname_output  = 'IEA-15-240-RWT_out.yaml'
    folder_output = os.getcwd() + os.sep + 'outputs'
    
    if not os.path.isdir(folder_output):
        os.mkdir(folder_output)

    Analysis_Level        = 0 # 0: Run CCBlade; 1: Update FAST model at each iteration but do not run; 2: Run FAST w/ ElastoDyn; 3: (Not implemented) Run FAST w/ BeamDyn
    # Initialize blade design
    refBlade = ReferenceBlade()
    refBlade.verbose  = True
    refBlade.NINPUT       = 8
    Nsection_Tow          = 12
    refBlade.NPTS         = 30
    refBlade.spar_var     = ['Spar_cap_ss', 'Spar_cap_ps'] # SS, then PS
    refBlade.te_var       = 'TE_reinforcement'
    refBlade.validate     = False
    refBlade.fname_schema = fname_schema
    blade = refBlade.initialize(fname_input)
    
    FASTpref                        = {}
    FASTpref['Analysis_Level']      = Analysis_Level
    # Set FAST Inputs
    if Analysis_Level >= 1:
        # File management
        FASTpref['FAST_ver']            = 'OpenFAST'
        FASTpref['dev_branch']          = True
        FASTpref['FAST_exe']            = '~/local/bin/openfast'
        FASTpref['FAST_directory']      = '../OpenFAST'   # Path to fst directory files
        FASTpref['FAST_InputFile']      = 'IEA-15-240-RWT.fst' # FAST input file (ext=.fst)
        FASTpref['Turbsim_exe']         = '~/local/bin/turbsim'
        FASTpref['FAST_namingOut']      = 'IEA-15-240-RWT'
        FASTpref['FAST_runDirectory']   = 'temp/' + FASTpref['FAST_namingOut']
        
        # Run Settings
        FASTpref['cores']               = 1
        FASTpref['debug_level']         = 2 # verbosity: set to 0 for quiet, 1 & 2 for increasing levels of output

        # DLCs
        FASTpref['DLC_gust']            = None      # Max deflection
        # FASTpref['DLC_gust']            = RotorSE_DLC_1_4_Rated       # Max deflection    ### Not in place yet
        FASTpref['DLC_extrm']           = None      # Max strain
        # FASTpref['DLC_extrm']           = RotorSE_DLC_7_1_Steady      # Max strain        ### Not in place yet
        FASTpref['DLC_turbulent']       = None
        # FASTpref['DLC_turbulent']       = RotorSE_DLC_1_1_Turb      # Alternate turbulent case, replacing rated and extreme DLCs for calculating max deflection and strain
        FASTpref['DLC_powercurve']      = None      # AEP
        # FASTpref['DLC_powercurve']      = None      # AEP

        # Initialize, read initial FAST files to avoid doing it iteratively
        fast = InputReader_OpenFAST(FAST_ver=FASTpref['FAST_ver'], dev_branch=FASTpref['dev_branch'])
        fast.FAST_InputFile = FASTpref['FAST_InputFile']
        fast.FAST_directory = FASTpref['FAST_directory']
        fast.execute()
        fst_vt = fast.fst_vt
    else:
        fst_vt = {}
    
    
    
    
    # Initialize and execute OpenMDAO problem with input data
    prob_ref = Problem()
    prob_ref.model=Optimize_MonopileTurbine(RefBlade=blade, Nsection_Tow = Nsection_Tow, VerbosityCosts = False, folder_output = folder_output, FASTpref = FASTpref)

    prob_ref.setup()
    prob_ref = Init_MonopileTurbine(prob_ref, blade, Nsection_Tow = Nsection_Tow, Analysis_Level = Analysis_Level, fst_vt = fst_vt)
    prob_ref['drive.shaft_angle']              = np.radians(6.)
    prob_ref['overhang']                       = 8.5
    prob_ref['drive.distance_hub2mb']          = 3.5
    prob_ref['significant_wave_height']        = 4.52
    prob_ref['significant_wave_period']        = 9.45
    prob_ref['monopile']                       = True
    prob_ref['foundation_height']              = -30.
    prob_ref['water_depth']                    = 30.
    prob_ref['suctionpile_depth']              = 45.
    prob_ref['wind_reference_height']          = 150.
    prob_ref['hub_height']                     = 150.
    prob_ref['tower_section_height']           = np.array([15., 15., 15., 15., 15., 15., 15., 15., 15., 15., 15., 10.])
    prob_ref['tower_outer_diameter']           = np.array([10., 10., 10., 9.999994, 9.893298, 9.501227, 9.073816, 8.733734, 8.481259, 8.254697, 8.087231, 7.512527, 6.717548])
    prob_ref['tower_wall_thickness']           = np.array([0.04922689, 0.04922689, 0.04922689, 0.04581482, 0.04301337, 0.04129422, 0.03939618, 0.03675472, 0.03345327, 0.02984231, 0.02622864, 0.03062863])
    # From SNOPT with better loading information:
    #prob_ref['tower_outer_diameter']           = np.array([10., 10., 9.07841266, 8.94626364, 8.2388643, 8.00406996, 7.4747645, 7.16117877, 6.7801444, 6.40652604, 6.14951031, 5.73067083, 5.57747088])
    #prob_ref['tower_wall_thickness']           = np.array([0.06470247, 0.05743258, 0.05558765, 0.05632417, 0.05654073, 0.05566434, 0.0556487, 0.05363243, 0.05230072, 0.05033409, 0.04682933, 0.04711726])

    prob_ref.model.nonlinear_solver = NonlinearRunOnce()
    prob_ref.model.linear_solver    = DirectSolver()
    print('Running at Initial Position:')
    prob_ref.run_driver()
    #prob_ref.model.list_inputs(units=True)
    #refBlade.write_ontology(fname_output, prob_ref['blade_out'], refBlade.wt_ref)
    
    print(prob_ref['hub_height'])
    print(prob_ref['rna_mass'])
    print('mIxx', prob_ref['tow.pre.mIxx'])
    print('mIyy', prob_ref['tow.pre.mIyy'])
    print('mIzz', prob_ref['tow.pre.mIzz'])
    print('mIxy', prob_ref['tow.pre.mIxy'])
    print('mIxz', prob_ref['tow.pre.mIxz'])
    print('mIyz', prob_ref['tow.pre.mIyz'])
    print('rna_F', prob_ref['tow.pre.rna_F'])
    print('rna_M', prob_ref['tow.pre.rna_M'])
    print('rna_cg', prob_ref['rna_cg'])
    print('Uref', prob_ref['tow.wind.Uref'])
    print('frequencies', prob_ref['tow.post.structural_frequencies'])
    print('stress', prob_ref['tow.post.stress'])
    print('local buckling', prob_ref['tow.post.shell_buckling'])
    print('shell buckling', prob_ref['tow.post.global_buckling'])

    #prob_ref.model.list_inputs(units=True)#values = False, hierarchical=False)
    #prob_ref.model.list_outputs(units=True)#values = False, hierarchical=False)    


    
    print('AEP =',                      prob_ref['AEP'])
    print('diameter =',                 prob_ref['diameter'])
    print('ratedConditions.V =',        prob_ref['rated_V'])
    print('ratedConditions.Omega =',    prob_ref['rated_Omega'])
    print('ratedConditions.pitch =',    prob_ref['rated_pitch'])
    print('ratedConditions.T =',        prob_ref['rated_T'])
    print('ratedConditions.Q =',        prob_ref['rated_Q'])
    print('mass_one_blade =',           prob_ref['mass_one_blade'])
    print('mass_all_blades =',          prob_ref['mass_all_blades'])
    print('I_all_blades =',             prob_ref['I_all_blades'])
    print('freq =',                     prob_ref['freq_pbeam'])
    print('tip_deflection =',           prob_ref['tip_deflection'])
    print('root_bending_moment =',      prob_ref['root_bending_moment'])
    print('moments at the hub =',       prob_ref['Mxyz_total'])
    print('blade cost =',               prob_ref['total_blade_cost'])

    # Run an optimization
    if optFlag:
        prob = Problem()
        prob.model=Optimize_MonopileTurbine(RefBlade=blade, Nsection_Tow = Nsection_Tow, folder_output = folder_output, user_update_routine=set_web3_offset)
        
        # --- Driver ---
        prob.driver = pyOptSparseDriver()
        prob.driver.options['optimizer'] = 'CONMIN'
        # prob.driver.opt_settings['ITMAX']     = 2
        # ----------------------

        # --- Objective ---
        prob.model.add_objective('lcoe')
        # ----------------------

        # --- Design Variables ---
        indices_no_root         = range(2,refBlade.NINPUT)
        indices_no_root_no_tip  = range(2,refBlade.NINPUT-1)
        prob.model.add_design_var('chord_in',    indices = indices_no_root_no_tip, lower=0.5,      upper=7.0)
        prob.model.add_design_var('theta_in',    indices = indices_no_root,        lower=-5.0,     upper=20.0)
        prob.model.add_design_var('sparT_in',    indices = indices_no_root_no_tip, lower=0.001,    upper=0.200)
        prob.model.add_design_var('control_tsr',                                   lower=6.000,    upper=11.00)
        # prob.model.add_design_var('tower_section_height', lower=5.0,  upper=80.0)
        # prob.model.add_design_var('tower_outer_diameter', lower=3.87, upper=10.0)
        # prob.model.add_design_var('tower_wall_thickness', lower=4e-3, upper=2e-1)
        # ----------------------
        
        # --- Constraints ---
        # Rotor
        prob.model.add_constraint('tip_deflection_ratio',     upper=1.0)  
        prob.model.add_constraint('AEP',                      lower=prob_ref['AEP'])
        # Tower
        # prob.model.add_constraint('tow.height_constraint',    lower=-1e-2,upper=1.e-2)
        # prob.model.add_constraint('tow.post.stress',          upper=1.0)
        # prob.model.add_constraint('tow.post.global_buckling', upper=1.0)
        # prob.model.add_constraint('tow.post.shell_buckling',  upper=1.0)
        # prob.model.add_constraint('tow.weldability',          upper=0.0)
        # prob.model.add_constraint('tow.manufacturability',    lower=0.0)
        # prob.model.add_constraint('frequencyNP_margin',       upper=0.)
        # prob.model.add_constraint('frequency1P_margin',       upper=0.)
        # prob.model.add_constraint('ground_clearance',         lower=20.0)
        # ----------------------
        
        # --- Recorder ---
        filename_opt_log = folder_output + os.sep + 'log_opt_' + blade['config']['name']
        
        prob.driver.add_recorder(SqliteRecorder(filename_opt_log))
        prob.driver.recording_options['includes'] = ['AEP','total_blade_cost','lcoe','tip_deflection_ratio']
        prob.driver.recording_options['record_objectives']  = True
        prob.driver.recording_options['record_constraints'] = True
        prob.driver.recording_options['record_desvars']     = True
        # ----------------------
        
        # --- Run ---
        prob.setup()
        prob = Init_MonopileTurbine(prob, blade, Nsection_Tow = Nsection_Tow)
        prob['drive.shaft_angle']              = np.radians(6.)
        prob['overhang']                       = 8.5
        prob['drive.distance_hub2mb']          = 3.5
        prob['significant_wave_height']        = 4.52
        prob['significant_wave_period']        = 9.45
        prob['monopile']                       = True
        prob['foundation_height']              = -30.
        prob['water_depth']                    = 30.
        prob['suctionpile_depth']              = 45.
        prob['wind_reference_height']          = 150.
        prob['hub_height']                     = 150.
        prob['tower_section_height']           = np.array([15., 15., 15., 15., 15., 15., 15., 15., 15., 15., 15., 10.])
        prob['tower_outer_diameter']           = np.array([10., 10., 10., 9.999994, 9.893298, 9.501227, 9.073816, 8.733734, 8.481259, 8.254697, 8.087231, 7.512527, 6.717548])
        prob['tower_wall_thickness']           = np.array([0.04922689, 0.04922689, 0.04922689, 0.04581482, 0.04301337, 0.04129422, 0.03939618, 0.03675472, 0.03345327, 0.02984231, 0.02622864, 0.03062863])
        # From SNOPT with better loading information:
        #prob['tower_outer_diameter']           = np.array([10., 10., 9.07841266, 8.94626364, 8.2388643, 8.00406996, 7.4747645, 7.16117877, 6.7801444, 6.40652604, 6.14951031, 5.73067083, 5.57747088])
        #prob['tower_wall_thickness']           = np.array([0.06470247, 0.05743258, 0.05558765, 0.05632417, 0.05654073, 0.05566434, 0.0556487, 0.05363243, 0.05230072, 0.05033409, 0.04682933, 0.04711726])
        prob.model.nonlinear_solver = NonlinearRunOnce()
        prob.model.linear_solver = DirectSolver()
        print('Running Optimization:')
        print('N design var: ', 2*len(indices_no_root_no_tip) + len(indices_no_root) + 1)
        prob.model.approx_totals()
        prob.run_driver()
        # ----------------------

        # --- Save output .yaml ---
        refBlade.write_ontology('optimization_out.yaml', prob['blade_out'], refBlade.wt_ref)
        shutil.copyfile(fname_input,  folder_output + os.sep + WT_input)

        # ----------------------
        # --- Outputs plotting ---
        print('AEP:         \t\t\t %f\t%f GWh \t Difference: %f %%' % (prob_ref['AEP']*1e-6, prob['AEP']*1e-6, (prob['AEP']-prob_ref['AEP'])/prob_ref['AEP']*100.))
        print('LCoE:        \t\t\t %f\t%f USD/MWh \t Difference: %f %%' % (prob_ref['lcoe']*1.e003, prob['lcoe']*1.e003, (prob['lcoe']-prob_ref['lcoe'])/prob_ref['lcoe']*100.))
        print('Blade cost:  \t\t\t %f\t%f USD \t Difference: %f %%' % (prob_ref['total_blade_cost'], prob['total_blade_cost'], (prob['total_blade_cost']-prob_ref['total_blade_cost'])/prob_ref['total_blade_cost']*100.))
        print('Blade mass:  \t\t\t %f\t%f kg  \t Difference: %f %%' % (prob_ref['total_blade_mass'], prob['total_blade_mass'], (prob['total_blade_mass']-prob_ref['total_blade_mass'])/prob_ref['total_blade_mass']*100.))
        print('Tower cost:  \t\t\t %f\t%f USD \t Difference: %f %%' % (prob_ref['tower_cost'], prob['tower_cost'], (prob['tower_cost']-prob_ref['tower_cost'])/prob_ref['tower_cost']*100.))
        print('Tower mass:  \t\t\t %f\t%f kg  \t Difference: %f %%' % (prob_ref['tower_mass'], prob['tower_mass'], (prob['tower_mass']-prob_ref['tower_mass'])/prob_ref['tower_mass']*100.))
        # ----------------------
    else:
        # If not optimizing, plot current design
        prob = prob_ref
        
    show_plots            = True
    flag_write_out        = True
    
    if flag_write_out:
        # --- Save output .yaml ---
        refBlade.write_ontology(fname_output, prob['blade_out'], refBlade.wt_ref)
        shutil.copyfile(fname_input,  folder_output + os.sep + fname_output)
    
    # Problem initialization
    var_y           = ['chord','theta','rthick','p_le','precurve','presweep']
    label_y         = ['Chord [m]', 'Twist [deg]', 'Relative Thickness [%]', 'Pitch Axis Location [%]', 'Prebend [m]', 'Sweep [m]']
    scaling_factor  = [1. , 1. , 100. , 100., 1., 1.]

    for i in range(len(var_y)):
        f1, ax1 = plt.subplots(1,1,figsize=(5.3, 4))
        ax1.plot(blade['pf']['r'], blade['pf'][var_y[i]] * scaling_factor[i])
        plt.xlabel('Rotor Coordinate [m]', fontsize=14, fontweight='bold')
        plt.ylabel(label_y[i], fontsize=14, fontweight='bold')
        plt.xticks(fontsize=12)
        plt.yticks(fontsize=12)
        plt.grid(color=[0.8,0.8,0.8], linestyle='--')
        plt.subplots_adjust(bottom = 0.15, left = 0.15)
        fig_name = var_y[i] + '_dimensional.png'
        f1.savefig(folder_output + os.sep + fig_name)

        f2, ax2 = plt.subplots(1,1,figsize=(5.3, 4))
        ax2.plot(blade['pf']['s'], blade['pf'][var_y[i]] * scaling_factor[i])
        plt.xlabel('Nondimensional Blade Span [-]', fontsize=14, fontweight='bold')
        plt.ylabel(label_y[i], fontsize=14, fontweight='bold')
        plt.xticks(fontsize=12)
        plt.yticks(fontsize=12)
        plt.grid(color=[0.8,0.8,0.8], linestyle='--')
        plt.subplots_adjust(bottom = 0.15, left = 0.15)
        fig_name = var_y[i] + '_nondimensional.png'
        f2.savefig(folder_output + os.sep + fig_name)

    # Pitch
    fp, axp  = plt.subplots(1,1,figsize=(5.3, 4))
    axp.plot(prob['V'], prob['pitch'])
    plt.xlabel('Wind [m/s]', fontsize=14, fontweight='bold')
    plt.ylabel('Pitch Angle [deg]', fontsize=14, fontweight='bold')
    #axp.legend(fontsize=12)
    plt.xticks(fontsize=12)
    plt.yticks(fontsize=12)
    plt.grid(color=[0.8,0.8,0.8], linestyle='--')
    plt.subplots_adjust(bottom = 0.15, left = 0.15)
    fig_name = 'pitch.png'
    fp.savefig(folder_output + os.sep + fig_name)

    # Power
    fpw, axpw  = plt.subplots(1,1,figsize=(5.3, 4))
    axpw.plot(prob['V'], prob['P'] * 1.00e-006)
    plt.xlabel('Wind [m/s]', fontsize=14, fontweight='bold')
    plt.ylabel('Electrical Power [MW]', fontsize=14, fontweight='bold')
    #axpw.legend(fontsize=12)
    plt.xticks(fontsize=12)
    plt.yticks(fontsize=12)
    plt.grid(color=[0.8,0.8,0.8], linestyle='--')
    plt.subplots_adjust(bottom = 0.15, left = 0.15)
    fig_name = 'power.png'
    fpw.savefig(folder_output + os.sep + fig_name)

    # Omega
    fo, axo  = plt.subplots(1,1,figsize=(5.3, 4))
    axo.plot(prob['V'], prob['Omega'])
    plt.xlabel('Wind [m/s]', fontsize=14, fontweight='bold')
    plt.ylabel('Rotor Speed [rpm]', fontsize=14, fontweight='bold')
    #axo.legend(fontsize=12)
    plt.xticks(fontsize=12)
    plt.yticks(fontsize=12)
    plt.grid(color=[0.8,0.8,0.8], linestyle='--')
    plt.subplots_adjust(bottom = 0.15, left = 0.15)
    fig_name = 'omega.png'
    fo.savefig(folder_output + os.sep + fig_name)

    # Tip speed
    fts, axts  = plt.subplots(1,1,figsize=(5.3, 4))
    axts.plot(prob['V'], prob['Omega'] * np.pi / 30. * prob['r'][-1])
    plt.xlabel('Wind [m/s]', fontsize=14, fontweight='bold')
    plt.ylabel('Blade Tip Speed [m/s]', fontsize=14, fontweight='bold')
    #axts.legend(fontsize=12)
    plt.xticks(fontsize=12)
    plt.yticks(fontsize=12)
    plt.grid(color=[0.8,0.8,0.8], linestyle='--')
    plt.subplots_adjust(bottom = 0.15, left = 0.15)
    fig_name = 'tip_speed.png'
    fts.savefig(folder_output + os.sep + fig_name)

    # Thrust
    ft, axt  = plt.subplots(1,1,figsize=(5.3, 4))
    axt.plot(prob['V'], prob['T'] * 1.00e-006)
    plt.xlabel('Wind [m/s]', fontsize=14, fontweight='bold')
    plt.ylabel('Rotor Thrust [MN]', fontsize=14, fontweight='bold')
    #axt.legend(fontsize=12)
    plt.xticks(fontsize=12)
    plt.yticks(fontsize=12)
    plt.grid(color=[0.8,0.8,0.8], linestyle='--')
    plt.subplots_adjust(bottom = 0.17, left = 0.15)
    fig_name = 'thrust.png'
    ft.savefig(folder_output + os.sep + fig_name)

    # Torque
    fq, axq  = plt.subplots(1,1,figsize=(5.3, 4))
    axq.plot(prob['V'], prob['Q'] * 1.00e-006)
    plt.xlabel('Wind [m/s]', fontsize=14, fontweight='bold')
    plt.ylabel('Rotor Torque [MNm]', fontsize=14, fontweight='bold')
    #axq.legend(fontsize=12)
    plt.xticks(fontsize=12)
    plt.yticks(fontsize=12)
    plt.grid(color=[0.8,0.8,0.8], linestyle='--')
    plt.subplots_adjust(bottom = 0.15, left = 0.15)
    fig_name = 'torque.png'
    fq.savefig(folder_output + os.sep + fig_name)

    # Tabular output
    temp = np.c_[blade['pf']['s'], blade['pf']['r']]
    for iy,y in enumerate(var_y):
        temp = np.c_[temp, blade['pf'][y]*scaling_factor[iy]]
    bladeDF = pd.DataFrame(data=temp, columns=['Blade Span','Rotor Coordinate [m]'] + label_y)
    
    perfDF = pd.DataFrame(data=np.c_[prob['V'],prob['pitch'], prob['P']*1e-6, prob['Omega'], prob['Omega']*prob['r'][-1]*np.pi/30., prob['T']*1e-6, prob['Q']*1e-6],
                         columns=['Wind [m/s]','Pitch [deg]','Power [MW]','Rotor Speed [rpm]','Tip Speed [m/s]','Thrust [MN]','Torque [MNm]'])

    wb = openpyxl.Workbook()
    sheetstr = ['Blade Design','Rotor Regulation Trajectory']
    for idf, df in enumerate([bladeDF, perfDF]):
        wb.create_sheet(sheetstr[idf])
        wb.active = wb.sheetnames.index(sheetstr[idf])
        ws = wb.active
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        for cell in ws["1:1"]:
            cell.style = 'Headline 2'
        ws.freeze_panes = ws['A2']

    # Remove empty sheet and write out
    wb.active = 0
    ws = wb.active
    wb.remove(ws)
    wb.save('IEA-15-240-RWT_rotor.xlsx')

    if show_plots == True:
        plt.show()
